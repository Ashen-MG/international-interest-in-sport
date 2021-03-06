from flask import request
from flasgger import SwaggerView, swag_from
from verification.jwt import is_admin
from openpyxl import load_workbook
from io import BytesIO
import csv_parsers.csvParser as parser
from csv_parsers.excelParser import excelParser
from settings import DB
import json

class UploadView(SwaggerView):

	@is_admin
	@swag_from("post.yml")
	def post(self):

		if len(request.files) == 0:
			return {"message": "Missing uploaded file."}, 400

		DB.createDatabaseBackup()

		fundingFile = request.files.get("fundingFile")
		successFile = request.files.get("successFile")
		interconnectednessFile = request.files.get("interconnectednessFile")
		bgsFile = request.files.get("bgsFile")

		requestJSON = json.loads(request.form["json"])
		countryCode = requestJSON.get("countryCode")
		currency = requestJSON.get("currency")
		# correction = requestJSON.get("correction")
		correction = []
		interconnectednessType = requestJSON.get("interconnectednessType")
		fundingSource = requestJSON.get("fundingSource")
		successSource = requestJSON.get("successSource")
		interconnSource = requestJSON.get("interconnSource")
		bgsSource = requestJSON.get("bgsSource")


		# At least one must be uploaded e.g. fundingFile could have been uploaded, but successFile and
		# interconnectednessFile are going to be None.

		if interconnectednessFile:
			wb = load_workbook(filename=BytesIO(interconnectednessFile.read()))

			p = excelParser()
			parsed = p.parseInterconnectness(wb, interconnectednessType)

			if DB.deleteInterconnectednessTables(interconnectednessType):
				for item in parsed:
					item.save()
				DB.saveNonFundingSource("interconnSource", interconnSource)
				#print(interconnSource)
			else:
				pass #TODO raise error alebo nieco ???

		if successFile:
			wb = load_workbook(filename=BytesIO(successFile.read()))

			p = excelParser()
			parsed = p.parseSuccess(wb)

			if DB.deleteSuccesTables():

				for item in parsed[0]:
					item.save()

				for table in parsed[2:]:
					table.save()

				#print(successSource)
				DB.saveNonFundingSource("successSource", successSource)

		"""	
		# TODO: list of unknown sports in parsed[1]
		"""

		if bgsFile:
			# wb = bgsFile
			wb = load_workbook(filename=BytesIO(bgsFile.read()))

			p = excelParser()
			parsed = p.parseBGS(wb)

			if DB.deleteBGS():

				for item in parsed:
					item.save()

				#print(bgsSource)
				DB.saveNonFundingSource("bgsSource", bgsSource)


		if fundingFile:
			# if correction is None:
				# return {"message": "Missing required parameter: `correction`.", "data": {}}, 400
			if not countryCode:
				return {"message": "Missing required parameter: `countryCode`.", "data": {}}, 400
			if not currency:
				return {"message": "Missing required parameter: `currency`.", "data": {}}, 400

			file = fundingFile
			lines = []
			for line in file:
				lines.append(line.decode("utf-8").strip())

			p = parser.csvParser()
			suggestions = p.findFailures(lines, correction, countryCode, currency)

			if len(suggestions) == 0:
				p.saveResults(countryCode)
				DB.saveFundingSource(countryCode, fundingSource)
			else:
				return {"message": "fail", "suggestions": suggestions}, 400

		return {"message": "ok"}